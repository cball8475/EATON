from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

ARIAL=lambda **k: Font(name="Arial", **k)
HDR_FILL=PatternFill("solid", fgColor="1F3864")
BAND=PatternFill("solid", fgColor="DCE6F1")
YEL=PatternFill("solid", fgColor="FFFF00")
THIN=Side(style="thin", color="BFBFBF")
BOX=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)

wb=Workbook(); ws=wb.active; ws.title="Roadmap"

ws["A1"]="EHS Project Roadmap — Sumter, SC"; ws["A1"].font=ARIAL(size=16,bold=True,color="1F3864")
ws["A2"]="Prepared by Charlie Ball for Kate Fowler · priorities 1-6 as assigned 31-Jul-2026 · targets run through Q2 2027"
ws["A2"].font=ARIAL(size=10,italic=True,color="595959")
ws["A3"]="HOW TO USE: Column I is yours — add milestones, dates or comments there. Columns A-H are maintained by Charlie; Status (F) updates at each 1:1. Yellow = your input."
ws["A3"].font=ARIAL(size=10,bold=True); ws["A3"].fill=YEL
ws["A4"]="Dates are working targets, not committed dates. Anything already past its date shows Status = Overdue. D1 Ref is the task or knowledge ID in the EHS command centre."
ws["A4"].font=ARIAL(size=9,italic=True,color="595959")

HEAD=["Pri","Workstream","Milestone / Deliverable","Owner","Target Date","Status","D1 Ref","Dependency / Note","Kate — Milestone or Comment"]
HR=6
for c,h in enumerate(HEAD,1):
    cell=ws.cell(HR,c,h); cell.font=ARIAL(size=10,bold=True,color="FFFFFF")
    cell.fill=HDR_FILL; cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=BOX

R=[
(1,"Respirators","Reschedule respirator inspection session","Charlie",date(2026,8,14),"In progress","#737","",""),
(1,"Respirators","Semi-annual respirator inspection & cleaning","Charlie",date(2026,9,30),"Not started","#470","Benchmark compliance item",""),
(1,"Respirators","Respirator TPM / cleaning schedule confirmed in place","Charlie",date(2026,9,30),"Not started","#626","",""),
(1,"Respirators","NIOSH 25/26/27 forms organised into folders","Charlie",date(2026,10,30),"Not started","#488","",""),
(1,"Respirators","Respiratory protection program annual review documented","Charlie",date(2026,11,30),"Not started","#523","Rides the MESH element review",""),
(1,"Respirators","2027 cycle scheduled + program closed out","Charlie",date(2027,1,29),"Not started","","",""),
(2,"Scrap Bin Standardization","Walk scrap process with Kate, Jad, Ashwin","Charlie",date(2026,8,21),"Not started","#525","Prerequisite to back-plate design",""),
(2,"Scrap Bin Standardization","Scrap-bin back plate designed","Charlie / Ashwin",date(2026,9,25),"Not started","#525","Bins rotate off 2nd rotator; plate removes manual dumping",""),
(2,"Scrap Bin Standardization","Back plate fabricated + trialled","ATS / Ashwin",date(2026,10,30),"Not started","#525","",""),
(2,"Scrap Bin Standardization","Gray / blue cart standard defined (5S+ shed)","Charlie",date(2026,10,30),"Not started","","",""),
(2,"Scrap Bin Standardization","Scrap-band standardization","Charlie",date(2026,11,30),"Not started","#524","One of the 5 roadmap risk-reduction projects",""),
(2,"Scrap Bin Standardization","Station cards issued for scrap handling","Charlie",date(2027,2,26),"Not started","#524","",""),
(2,"Scrap Bin Standardization","Standard rolled to all fab areas","Charlie",date(2027,3,31),"Not started","","",""),
(3,"Lifting Devices","Vacuum lift solenoid project complete + GHG updated","Charlie",date(2026,8,14),"Overdue","#691","Benchmark #162, closure was 31-Jul-2026",""),
(3,"Lifting Devices","Crane inspections — Engineered Systems Inc","Kate",date(2026,8,31),"In progress","#616","Kate gathering cost for corporate across sites",""),
(3,"Lifting Devices","Q3 CapEx ~$250k received","Kate",date(2026,8,31),"In progress","KN #467","Budget gate for every purchase below",""),
(3,"Lifting Devices","Crane copper lifting devices — implementation plan","Charlie",date(2026,8,28),"In progress","#696","",""),
(3,"Lifting Devices","Replacements specified + quoted","Charlie / Kate",date(2026,9,18),"Not started","KN #467","Sheet lifter by paint line + 3 others + big-door lifter lasers to conveyors",""),
(3,"Lifting Devices","PO placed / 10% down","Kate",date(2026,9,30),"Not started","","10% down locks the project permanently",""),
(3,"Lifting Devices","Crane inspection checklist verified for under-hook devices","Charlie",date(2026,9,30),"Not started","#309","Badger lesson learned — 27-stitch laceration",""),
(3,"Lifting Devices","Overhead crane PM ownership resolved","Charlie / Kate / Gireesh",date(2026,10,30),"Not started","KN #442","Disputed between supply chain and fab; no PM owner",""),
(3,"Lifting Devices","Setup-frame crane retired, A-frame replaced","Kate",date(2026,11,27),"Not started","KN #407","IES spare motor + parts; Hodges shipment held",""),
(3,"Lifting Devices","Yellow gym-hoist training documented in setup","Charlie / Marcus",date(2026,11,27),"Not started","#618","",""),
(3,"Lifting Devices","New lifting devices delivered + installed","Charlie",date(2027,3,31),"Not started","","Lead time dependent on Sep PO",""),
(3,"Lifting Devices","Post-install verification + WSRA updates","Charlie",date(2027,4,30),"Not started","","",""),
(4,"Press Brakes","Press-brake safety alert (proper tool use) sent + confirmed","Charlie",date(2026,8,7),"In progress","#687","",""),
(4,"Press Brakes","Dept-wide comm: tools, or no hands on press brakes","Charlie",date(2026,8,7),"Overdue","#561 / #653","",""),
(4,"Press Brakes","Operator safety talk-down run","Charlie",date(2026,8,14),"Overdue","#654","",""),
(4,"Press Brakes","Formal near-miss record opened + Ashwin meeting","Charlie / Ashwin",date(2026,8,14),"Not started","#623","",""),
(4,"Press Brakes","Method sheet updated for changing dies","Charlie",date(2026,8,21),"Overdue","#690","",""),
(4,"Press Brakes","High-risk part population defined + safe routing method","Charlie / Ashwin",date(2026,9,11),"Not started","#664","Gates the routing milestone below",""),
(4,"Press Brakes","Press-brake training package built","Charlie",date(2026,9,30),"Not started","#660","Ron Brown / Marty Honey",""),
(4,"Press Brakes","AI camera / vision system specified + ordered","Charlie / Ashwin",date(2026,10,30),"Not started","#521","Needs Q3 CapEx confirmed",""),
(4,"Press Brakes","HRPs routed through the 4 safe press brakes + Amada","Charlie / Ashwin",date(2026,11,30),"Not started","#521","See Assumptions tab — definition of done",""),
(4,"Press Brakes","New-equipment safety standard written","Charlie",date(2026,11,30),"Not started","#521","Sumter leads this corporately",""),
(4,"Press Brakes","Weekly press-brake safety meeting","Charlie",date(2026,12,31),"In progress","#658","Recurring, runs through the project",""),
(4,"Press Brakes","AI camera system installed + validated","Charlie / Ashwin",date(2027,3,31),"Not started","#521","",""),
(5,"Laceration Risks","May lessons learned entered","Charlie",date(2026,8,14),"Overdue","#545","",""),
(5,"Laceration Risks","Finger-laceration reportable + corrective actions distributed","Kate",date(2026,8,14),"Not started","#444","",""),
(5,"Laceration Risks","Offal staging carts with protruding pieces eliminated","Charlie",date(2026,9,25),"In progress","#515","",""),
(5,"Laceration Risks","L-bracket / guard under conveyor machine return","Charlie",date(2026,9,25),"Not started","#336","Also addresses heat reflection",""),
(5,"Laceration Risks","Band-cutting standard work closed","Charlie",date(2026,10,30),"Not started","#632","Falling-channels near-miss",""),
(5,"Laceration Risks","Offal racks — ceiling, height control, accountability","Charlie",date(2026,11,30),"Not started","#524","One of the 5 roadmap risk-reduction projects",""),
(5,"Laceration Risks","Pallet-stacker procedural review + rack-lift procedure","Charlie / David",date(2026,11,30),"Not started","#688","Triggered on new stacker arrival",""),
(5,"Laceration Risks","Laceration risk register consolidated (bins, carts, conveyors, pallets)","Charlie",date(2026,12,18),"Not started","","Covers 4 prior recordables",""),
(5,"Laceration Risks","2027 verification pass + WSRA integration","Charlie",date(2027,3,31),"Not started","","",""),
(6,"Laser Loading","Trumpf contacted re OEM loading method","Charlie",date(2026,8,14),"In progress","#554","Roller balls up or down; is the clamshell the right device",""),
(6,"Laser Loading","Ashwin reviews Laser 1 loading mod for replication","Ashwin",date(2026,8,28),"Not started","#554","",""),
(6,"Laser Loading","Scope confirmed against 2027 laser divestiture","Charlie / Kate",date(2026,9,30),"Not started","KN #416","See Assumptions tab — two suppliers buying the lasers",""),
(6,"Laser Loading","Laser 3 near-miss corrective actions closed","Charlie",date(2026,9,25),"Not started","KN #441","Sheet slid, creased 480V control-panel line",""),
(6,"Laser Loading","Laser 2 & 3 loading mods implemented","Ashwin / ATS",date(2026,11,30),"Not started","#554","Conditional on divestiture scope above",""),
(6,"Laser Loading","Overhead crane maintenance / PM restored","Kate / Gireesh",date(2026,12,18),"Not started","KN #442","~10 yrs unmaintained; single point of failure for laser loading",""),
]
r=HR+1
for row in R:
    for c,v in enumerate(row,1):
        cell=ws.cell(r,c,v); cell.font=ARIAL(size=10); cell.border=BOX
        cell.alignment=Alignment(vertical="top",wrap_text=(c in (3,8,9)))
        if c==5: cell.number_format="dd-mmm-yyyy"; cell.alignment=Alignment(horizontal="center",vertical="top")
        if c in (1,7): cell.alignment=Alignment(horizontal="center",vertical="top")
        if c==6:
            cell.alignment=Alignment(horizontal="center",vertical="top")
            if v=="Overdue": cell.font=ARIAL(size=10,bold=True,color="C00000")
            elif v=="In progress": cell.font=ARIAL(size=10,color="1F6F3D")
        if c==9: cell.fill=YEL
    if row[0]%2==0:
        for c in range(1,9): ws.cell(r,c).fill=BAND
    r+=1
LAST=r-1
ws.cell(HR+1,9,"e.g. Gate review with Laura — 15-Sep").font=ARIAL(size=9,italic=True,color="808080")

for col,w in zip("ABCDEFGHI",[5,24,52,22,14,13,12,46,34]):
    ws.column_dimensions[col].width=w
ws.row_dimensions[HR].height=30
ws.freeze_panes="C7"
ws.auto_filter.ref=f"A{HR}:I{LAST}"

# ---- Summary ----
s=wb.create_sheet("Summary")
s["A1"]="Roadmap Summary"; s["A1"].font=ARIAL(size=14,bold=True,color="1F3864")
s["A2"]="All figures calculated from the Roadmap tab — they update as Status changes."
s["A2"].font=ARIAL(size=9,italic=True,color="595959")
sh=["Pri","Workstream","Milestones","Complete","In progress","Overdue","Not started","% Complete","Last target date"]
for c,h in enumerate(sh,1):
    cell=s.cell(4,c,h); cell.font=ARIAL(size=10,bold=True,color="FFFFFF"); cell.fill=HDR_FILL
    cell.alignment=Alignment(horizontal="center",wrap_text=True); cell.border=BOX
WS=[(1,"Respirators"),(2,"Scrap Bin Standardization"),(3,"Lifting Devices"),(4,"Press Brakes"),(5,"Laceration Risks"),(6,"Laser Loading")]
# rows are grouped contiguously by priority -> use block ranges (plain Excel 2007 funcs only)
blocks={}
for i,row in enumerate(R):
    blocks.setdefault(row[0],[]).append(HR+1+i)
for i,(p,name) in enumerate(WS):
    rr=5+i; a0,a1=min(blocks[p]),max(blocks[p])
    st=f"Roadmap!$F${a0}:$F${a1}"; dt=f"Roadmap!$E${a0}:$E${a1}"
    s.cell(rr,1,p).alignment=Alignment(horizontal="center")
    s.cell(rr,2,name)
    s.cell(rr,3,a1-a0+1)
    s.cell(rr,4,f'=COUNTIF({st},"Complete")')
    s.cell(rr,5,f'=COUNTIF({st},"In progress")')
    s.cell(rr,6,f'=COUNTIF({st},"Overdue")')
    s.cell(rr,7,f'=COUNTIF({st},"Not started")')
    s.cell(rr,8,f'=IF($C{rr}=0,0,$D{rr}/$C{rr})')
    s.cell(rr,9,f'=MAX({dt})')
    for c in range(1,10):
        cell=s.cell(rr,c); cell.font=ARIAL(size=10); cell.border=BOX
        if c==8: cell.number_format="0.0%"
        if c==9: cell.number_format="dd-mmm-yyyy"
        if c in (3,4,5,6,7): cell.alignment=Alignment(horizontal="center")
TR=5+len(WS)
s.cell(TR,2,"TOTAL").font=ARIAL(size=10,bold=True)
for c,L in zip(range(3,8),"CDEFG"):
    s.cell(TR,c,f'=SUM({L}5:{L}{TR-1})')
s.cell(TR,8,f'=IF($C{TR}=0,0,$D{TR}/$C{TR})')
s.cell(TR,9,f'=MAX(Roadmap!$E${HR+1}:$E${LAST})')
for c in range(1,10):
    cell=s.cell(TR,c); cell.font=ARIAL(size=10,bold=True); cell.border=BOX
    cell.fill=PatternFill("solid",fgColor="DCE6F1")
    if c==8: cell.number_format="0.0%"
    if c==9: cell.number_format="dd-mmm-yyyy"
    if c in (3,4,5,6,7): cell.alignment=Alignment(horizontal="center")
s.cell(TR+2,2,"Counts read Status on the Roadmap tab; % Complete = Complete / Milestones. Update Status there and these refresh.").font=ARIAL(size=9,italic=True,color="595959")
for col,w in zip("ABCDEFGHI",[5,26,11,11,12,10,12,12,16]): s.column_dimensions[col].width=w
s.row_dimensions[4].height=28

# ---- Assumptions ----
a=wb.create_sheet("Assumptions & Open Items")
a["A1"]="Assumptions and Open Items"; a["A1"].font=ARIAL(size=14,bold=True,color="1F3864")
a["A2"]="These drive the dates on the Roadmap tab. If an assumption is wrong the dates move — flag it and I will reissue."
a["A2"].font=ARIAL(size=9,italic=True,color="595959")
for c,h in enumerate(["#","Item","Assumption made","What I need from you"],1):
    cell=a.cell(4,c,h); cell.font=ARIAL(size=10,bold=True,color="FFFFFF"); cell.fill=HDR_FILL
    cell.alignment=Alignment(horizontal="center",wrap_text=True); cell.border=BOX
AS=[
(1,"Press brakes — definition of done","I have taken 'done' to mean all three: high-risk parts identified and routed through the 4 safe press brakes and the Amada, the AI camera system installed and validated, and the new-equipment standard written. Routing and the standard land in 2026; camera install runs into Q1 2027 on lead time.","Confirm this is the finish line you have in mind, or tell me which of the three you are dating."),
(2,"Laser loading vs 2027 divestiture","Two suppliers are buying the lasers, punches and spot welders as Sumter moves to switchboard-only in 2027 (KN #416). I have scoped Laser 2 and 3 loading mods as conditional and put a scope-confirmation gate in September before any spend.","Whether we invest in loading mods for assets that leave in 2027, or hold at controls and training only."),
(3,"Q3 CapEx gate","Lifting device purchases and the AI camera order both assume the ~$250k Q3 CapEx lands in August as planned. PO placed September with 10% down to lock it.","Confirm the August timing and whether the AI camera comes out of this pot or a separate one."),
(4,"Respirators scope","Treated as a compliance-maintenance workstream, not a build — inspection cadence, TPM schedule, forms and the annual program review. No new respirator selection or fit-test program assumed.","Confirm nothing larger is intended here."),
(5,"Owner assignments","Where a task had no owner in the system I have put myself. Ashwin carries the engineering reviews, ATS the fabrication and install.","Correct anything I have assigned to the wrong person."),
(6,"Overdue items","Five items were already past their date at issue and are dated forward into August: vacuum lift solenoid (Benchmark #162), the two press-brake communications, the operator talk-down and the die-change method sheet.","No action needed — flagging so the red on the sheet is not a surprise."),
(7,"Dates are targets","Working targets for sequencing and 1:1 review, not committed dates to corporate. Nothing here has been shared outside the two of us.","Tell me which of these you want turned into committed dates."),
]
r=5
for row in AS:
    for c,v in enumerate(row,1):
        cell=a.cell(r,c,v); cell.font=ARIAL(size=10); cell.border=BOX
        cell.alignment=Alignment(vertical="top",wrap_text=(c in (2,3,4)))
        if c==1: cell.alignment=Alignment(horizontal="center",vertical="top")
        if c==4: cell.fill=YEL
    r+=1
for col,w in zip("ABCD",[5,30,72,50]): a.column_dimensions[col].width=w
a.row_dimensions[4].height=28

wb.save("EHS_Project_Roadmap_2026-2027.xlsx")
print("written")
